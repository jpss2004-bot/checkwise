"""Phase 10 — server-side report export pipeline.

Walks ``ReportVersion.content_json``, renders the block list into a
self-contained artifact, writes the artifact to storage, and updates
``ReportExport.status`` along the way. The state machine the model
docstring promised (pending → rendering → ready | failed) finally
gets a worker.

Slice 10A ships HTML rendering (pure Python, no deps). Slice 10B
adds PDF rendering by loading the 10A HTML in headless Chromium via
Playwright and calling page.pdf(). Slice 10C (Excel) and 10D (signed
sharing) reuse the same dispatcher and storage path — the format
dispatch lives in :func:`run_report_export`.

Design choices:

* **No new dependency.** HTML assembly is done in pure Python via
  ``html.escape`` plus a small set of helper functions. Jinja2 isn't
  installed and adding it for one templated artifact is overkill;
  10B will likely shell out to Playwright rendering the existing
  React print page rather than templating server-side anyway.

* **Generic block fallback.** The 10 block types each have a custom
  React Canvas component in the editor. Reimplementing each layout
  in server-side HTML is out of scope for the foundation slice
  (10B's Playwright path will reuse the React layouts). Instead the
  renderer takes a structural approach: title + scope, then for
  every block a card with the block type, configured title, and a
  recursive key/value dump of ``block.data``. Reads cleanly, lossless
  for downstream consumers (LLM ingestion, raw data review), and
  doesn't lie about the visual fidelity 10B will deliver.

* **The worker never raises.** ``run_report_export`` catches every
  exception and writes ``status="failed"`` with ``error_text`` set
  to the message. Callers (currently a FastAPI ``BackgroundTask``)
  see no exception — the failure is visible only via the polling
  endpoint. Future async runners get the same contract for free.

* **Storage abstraction.** Reuses :class:`StorageService.save_bytes`,
  the same path the document-upload pipeline uses. Local + S3
  backends both work out of the box.

Out of scope here:
    * PDF / Excel / DOCX rendering (Slices 10B, 10C).
    * Signed-link sharing for unauthenticated viewers (Slice 10D).
    * Watermarks, password-protected exports (Slice 10D).
    * Email-the-PDF flows (depends on the deferred Phase 6 email work).
"""

from __future__ import annotations

import html
import logging
import tempfile
from collections.abc import Iterable
from pathlib import Path
from typing import Any, Literal

from sqlalchemy.orm import Session

from app.constants.reports import ReportAudience
from app.models import Report, ReportExport, ReportVersion
from app.models.entities import new_id, utc_now
from app.services.storage import get_storage_service

logger = logging.getLogger(__name__)


ReportExportFormat = Literal["html", "pdf", "xlsx"]
SUPPORTED_FORMATS: tuple[ReportExportFormat, ...] = ("html", "pdf", "xlsx")


class ReportExportError(Exception):
    """Raised by ``start_report_export`` on bad input.

    ``run_report_export`` *never* raises — it catches everything and
    persists ``status="failed"``. The start helper raises so the API
    layer can map to 422 / 404 cleanly.
    """


# ---------------------------------------------------------------------------
# State-machine entry points
# ---------------------------------------------------------------------------


def start_report_export(
    db: Session,
    *,
    report: Report,
    version: ReportVersion,
    format: str,
    requested_by_user_id: str,
) -> ReportExport:
    """Create a pending ``ReportExport`` row.

    Validates the format before persisting. Caller is responsible for
    scheduling :func:`run_report_export` (typically via a FastAPI
    ``BackgroundTask`` so the request returns immediately).

    Does NOT commit — the caller owns the transaction so the
    enqueueing of the background task can be tied to the commit
    boundary (we don't want to schedule a job that references a row
    that never lands).
    """
    if format not in SUPPORTED_FORMATS:
        raise ReportExportError(
            f"Unsupported export format '{format}'. Supported: "
            f"{', '.join(SUPPORTED_FORMATS)}."
        )
    row = ReportExport(
        report_id=report.id,
        version_id=version.id,
        format=format,
        status="pending",
        requested_by_user_id=requested_by_user_id,
    )
    db.add(row)
    db.flush()
    return row


def run_report_export(db: Session, export_id: str) -> None:
    """Render the export's artifact and update the row to ``ready``.

    Never raises. On any failure the row is updated with
    ``status="failed"`` and the exception message in ``error_text``,
    so the polling endpoint can show the user what went wrong.

    The caller is responsible for the surrounding session lifecycle.
    In a FastAPI ``BackgroundTask`` context that's typically a fresh
    ``SessionLocal()`` per task — the request session is already
    closed by the time the background task runs.
    """
    row = db.get(ReportExport, export_id)
    if row is None:
        logger.error("[reports.export] export %s not found, abandoning", export_id)
        return

    try:
        row.status = "rendering"
        db.flush()

        report = db.get(Report, row.report_id)
        version = db.get(ReportVersion, row.version_id)
        if report is None or version is None:
            raise ReportExportError(
                "Report or version was deleted before the export ran."
            )

        if row.format == "html":
            artifact = render_report_html(report, version)
            content_type = "text/html; charset=utf-8"
            extension = "html"
        elif row.format == "pdf":
            artifact = render_report_pdf(report, version)
            content_type = "application/pdf"
            extension = "pdf"
        elif row.format == "xlsx":
            artifact = render_report_xlsx(report, version)
            content_type = (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            extension = "xlsx"
        else:
            # Defensive — start_report_export already validated, but a
            # stale row from a prior schema could carry an unsupported
            # value. Fail cleanly.
            raise ReportExportError(
                f"Renderer not wired for format '{row.format}'."
            )

        storage = get_storage_service()
        # Deterministic-ish key with a random suffix so retries don't
        # collide on the same name. Subdir scopes per report for both
        # ops audit (``s3://bucket/report-exports/<report_id>/``) and
        # bulk-delete-on-report-deletion in a future cleanup slice.
        storage_key = (
            f"report-exports/{report.id}/"
            f"v{version.version_number}-{new_id()}.{extension}"
        )
        storage.save_bytes(
            storage_key=storage_key,
            data=artifact,
            content_type=content_type,
        )

        row.storage_key = storage_key
        row.bytes = len(artifact)
        row.status = "ready"
        row.ready_at = utc_now()
        row.error_text = None
        db.flush()
    except Exception as exc:  # noqa: BLE001 — see docstring; never re-raise
        logger.exception(
            "[reports.export] render failed for export_id=%s", export_id
        )
        try:
            row.status = "failed"
            row.error_text = str(exc)[:1000]  # truncate to the column's text capacity
            db.flush()
        except Exception:  # pragma: no cover — defensive double-failure path
            logger.exception(
                "[reports.export] failed to persist failure state for %s", export_id
            )


# ---------------------------------------------------------------------------
# Pure rendering (no DB, no I/O)
# ---------------------------------------------------------------------------


def render_report_pdf(report: Report, version: ReportVersion) -> bytes:
    """Phase 10B — render the 10A HTML artifact through headless Chromium.

    Approach: render the same self-contained HTML the 10A path produces,
    write it to a temporary ``.html`` file, load it via ``file://`` in a
    headless Chromium instance, call ``page.pdf()``. Browser launches
    fresh per call and closes immediately — no pooling. RAM stays
    bounded at ~150-300MB peak per render so we can survive on Render's
    starter plan (512MB) under normal traffic.

    Why a temp file (not a data URL or in-memory content): the artifact
    is small enough that the temp-file write is sub-millisecond, AND
    Chromium's ``file://`` loader handles the resulting page exactly
    like a real browser would handle a saved HTML — same media query
    layout, same print stylesheet, same fonts. Data URLs cap out around
    2MB on some Chromium versions and don't trigger the same media
    handling for embedded ``<style media=print>`` rules.

    PDF settings: US Letter, 0.5in margins all around, ``print_background=True``
    so the inline CSS gradients / background colors render. These are
    the conservative defaults that match the existing browser-print
    flow. Future slices can expose them as request parameters.

    Lifecycle: the Playwright sync context manager guarantees the
    browser closes even if ``page.pdf`` raises. The temp file is
    deleted after read via ``Path.unlink`` in the finally — no orphan
    files in /tmp under repeated failure modes.

    Import is local because ``playwright`` is a heavy import that adds
    ~150ms to module-load time. The 10A path doesn't pay that cost for
    the HTML-only consumer.
    """
    from playwright.sync_api import sync_playwright

    from app.services.reports.print_render import render_report_document_html

    # The PDF renders the DESIGNED document (verdict → findings → bars →
    # matrix), not the generic key/value dump that render_report_html still
    # produces for the raw-data HTML export.
    html_bytes = render_report_document_html(report, version)
    # ``delete=False`` so the temp file survives the ``with`` block —
    # we delete it explicitly in the outer ``finally`` so Chromium has
    # the file path available throughout the render.
    tmp = tempfile.NamedTemporaryFile(
        suffix=".html", delete=False, prefix="cw-report-"
    )
    try:
        tmp.write(html_bytes)
        tmp.close()
        tmp_path = Path(tmp.name)
        with sync_playwright() as p:
            browser = p.chromium.launch(args=["--no-sandbox"])
            try:
                page = browser.new_page()
                page.goto(f"file://{tmp_path}", wait_until="networkidle")
                pdf_bytes = page.pdf(
                    format="Letter",
                    print_background=True,
                    margin={
                        "top": "0.5in",
                        "right": "0.5in",
                        "bottom": "0.5in",
                        "left": "0.5in",
                    },
                )
            finally:
                browser.close()
        return pdf_bytes
    finally:
        try:
            Path(tmp.name).unlink(missing_ok=True)
        except Exception:  # pragma: no cover — defensive cleanup
            logger.warning(
                "[reports.export] failed to delete temp file %s", tmp.name
            )


def render_report_xlsx(report: Report, version: ReportVersion) -> bytes:
    """Phase 10C — render the report as an Excel workbook.

    Layout: one "Cover" sheet with the report metadata, then one
    sheet per block. Each block sheet carries the block's title in
    A1 and a 2-column key/value dump of ``block.data`` starting in
    A3. Tabular block payloads (lists of dicts — the typical
    vendor_risk_matrix shape) get flattened into a proper header
    row + value rows so the user can sort / filter / pivot like any
    other Excel table.

    Sheet names are Excel-safe: trimmed to 31 chars, with the
    illegal chars ``[]:*?/\\`` replaced by underscores, and a
    numeric suffix on collisions. Empty/divider blocks are skipped
    (they don't add anything as a separate sheet).

    Pure Python; no native deps. The output bytes are written via
    ``Workbook.save`` to a ``BytesIO`` so the dispatcher never
    touches the filesystem — slight win over the PDF path's
    tempfile dance.

    Import is local so the 10A/10B paths don't pay the openpyxl
    module-load cost.
    """
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    cover = wb.active
    cover.title = "Resumen"

    # Cover sheet ─────────────────────────────────────────────────
    cover["A1"] = report.title or "Reporte CheckWise"
    cover["A1"].font = Font(size=18, bold=True)
    if report.description:
        cover["A2"] = report.description
        cover["A2"].alignment = Alignment(wrap_text=True)
    cover["A4"] = "Audiencia"
    cover["B4"] = _audience_label(report.audience)
    cover["A5"] = "Versión"
    cover["B5"] = f"v{version.version_number}"
    cover["A6"] = "Generado"
    cover["B6"] = utc_now().isoformat()
    blocks = (version.content_json or {}).get("blocks") or []
    cover["A7"] = "Bloques"
    cover["B7"] = len(blocks)
    for row_idx in range(4, 8):
        cover.cell(row=row_idx, column=1).font = Font(bold=True)
    cover.column_dimensions["A"].width = 16
    cover.column_dimensions["B"].width = 60

    # Per-block sheets ────────────────────────────────────────────
    used_names: set[str] = {cover.title}
    header_fill = PatternFill(
        start_color="FF013557", end_color="FF013557", fill_type="solid"
    )
    header_font = Font(color="FFFFFFFF", bold=True)

    for index, block in enumerate(blocks, start=1):
        block_type = str(block.get("type") or "unknown")
        if block_type == "divider":
            continue
        config = block.get("config") or {}
        data = block.get("data") or {}
        title = str(config.get("title") or _humanise_type(block_type))
        sheet_name = _safe_sheet_name(f"{index:02d} {title}", used_names)
        used_names.add(sheet_name)
        ws = wb.create_sheet(sheet_name)

        # Title row
        ws["A1"] = title
        ws["A1"].font = Font(size=14, bold=True)
        ws["A2"] = f"type: {block_type}"
        ws["A2"].font = Font(italic=True, color="FF666666")

        if block_type == "text":
            body = str(config.get("body") or data.get("body") or "")
            ws["A4"] = body
            ws["A4"].alignment = Alignment(wrap_text=True, vertical="top")
            ws.column_dimensions["A"].width = 100
            continue

        # Try the tabular shape first — a list of dicts under any
        # top-level key renders as a real table. ``vendor_risk_matrix``
        # is the canonical example; ``kpi_strip.resolved`` and
        # ``attention_list.items`` follow the same shape.
        rendered_tabular = _try_render_tabular_block(
            ws, data, header_fill=header_fill, header_font=header_font
        )
        if rendered_tabular:
            continue

        # Generic key/value dump otherwise.
        _render_keyvalue_block(ws, data)
        ws.column_dimensions["A"].width = 32
        ws.column_dimensions["B"].width = 80

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _safe_sheet_name(name: str, used: set[str]) -> str:
    """Coerce ``name`` into an Excel-legal sheet name, deduping on collision.

    Excel sheet names: max 31 chars, no ``[]:*?/\\``, can't be empty,
    can't collide. We replace illegal chars with ``_`` and append
    ``-2``, ``-3``, … to disambiguate.
    """
    cleaned = "".join("_" if c in "[]:*?/\\" else c for c in name).strip() or "Bloque"
    truncated = cleaned[:31]
    if truncated not in used:
        return truncated
    base = truncated[:28]  # leave room for "-NN"
    for i in range(2, 100):
        candidate = f"{base}-{i}"
        if candidate not in used:
            return candidate
    return truncated[:31]  # last-resort fallback — caller will warn on dupe


def _try_render_tabular_block(
    ws,
    data: Any,
    *,
    header_fill,
    header_font,
) -> bool:
    """Look for a list of dicts inside ``data`` and render it as a table.

    Returns True if it rendered, False otherwise. The first top-level
    value that's a non-empty list of dicts wins — most block payloads
    only carry one tabular surface (``vendor_risk_matrix.vendors``,
    ``kpi_strip.resolved``, etc).
    """
    from openpyxl.styles import Font

    if not isinstance(data, dict):
        return False
    for key, value in data.items():
        if (
            isinstance(value, list)
            and value
            and all(isinstance(item, dict) for item in value)
        ):
            ws["A4"] = key
            ws["A4"].font = Font(bold=True, color="FF02558A")
            # Stable header order: union of keys, sorted, with common
            # identity-looking keys first.
            all_keys: list[str] = []
            seen: set[str] = set()
            for item in value:
                for k in item.keys():
                    if k not in seen:
                        seen.add(k)
                        all_keys.append(str(k))
            for col_idx, header in enumerate(all_keys, start=1):
                cell = ws.cell(row=6, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
            for row_offset, item in enumerate(_truncate_iter(value), start=7):
                if not isinstance(item, dict):
                    continue
                for col_idx, header in enumerate(all_keys, start=1):
                    cell_val = item.get(header)
                    if isinstance(cell_val, (dict, list, tuple)):
                        cell_val = str(cell_val)
                    elif cell_val is None:
                        cell_val = ""
                    ws.cell(row=row_offset, column=col_idx, value=cell_val)
            for col_idx in range(1, len(all_keys) + 1):
                ws.column_dimensions[
                    ws.cell(row=6, column=col_idx).column_letter
                ].width = 22
            return True
    return False


def _render_keyvalue_block(ws, data: Any, *, start_row: int = 4) -> None:
    """Recursive key/value dump into rows 4+ of ``ws``.

    Dict → one row per key with the value stringified (nested dicts
    flatten to ``json.dumps``-style strings). Non-dict payloads end
    up in a single cell so the sheet isn't blank.
    """
    from openpyxl.styles import Alignment, Font

    if not isinstance(data, dict):
        ws["A4"] = str(data) if data is not None else "(sin datos)"
        return
    row = start_row
    for key, value in data.items():
        ws.cell(row=row, column=1, value=str(key)).font = Font(bold=True)
        rendered: str
        if value is None:
            rendered = "—"
        elif isinstance(value, bool):
            rendered = "Sí" if value else "No"
        elif isinstance(value, (str, int, float)):
            rendered = str(value)
        else:
            # Fall back to a compact JSON-ish repr for nested
            # structures so the cell stays readable.
            import json as _json

            try:
                rendered = _json.dumps(value, ensure_ascii=False, default=str)[:1000]
            except Exception:
                rendered = str(value)[:1000]
        cell = ws.cell(row=row, column=2, value=rendered)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        row += 1


def render_report_html(report: Report, version: ReportVersion) -> bytes:
    """Render a report version into a self-contained HTML document.

    Output is UTF-8 bytes. No external assets are referenced. All
    styling is inlined in a single ``<style>`` block. The document
    is safe to open in any browser or to attach to an email.
    """
    content = version.content_json or {}
    blocks = content.get("blocks") or []
    audience = _audience_label(report.audience)
    title = html.escape(report.title or "Reporte CheckWise")
    description = html.escape(report.description or "")
    generated_at = utc_now().isoformat()

    parts: list[str] = []
    parts.append("<!doctype html>")
    parts.append("<html lang=\"es\">")
    parts.append("<head>")
    parts.append("<meta charset=\"utf-8\">")
    parts.append(
        "<meta name=\"viewport\" content=\"width=device-width, initial-scale=1\">"
    )
    parts.append(f"<title>{title} · CheckWise</title>")
    parts.append(_STYLE_BLOCK)
    parts.append("</head>")
    parts.append("<body>")
    parts.append("<main class=\"page\">")

    parts.append("<header class=\"cover\">")
    parts.append("<p class=\"eyebrow\">CheckWise · Reporte exportado</p>")
    parts.append(f"<h1>{title}</h1>")
    if description:
        parts.append(f"<p class=\"description\">{description}</p>")
    parts.append("<dl class=\"cover-meta\">")
    parts.append(_meta_pair("Audiencia", html.escape(audience)))
    parts.append(_meta_pair("Versión", f"v{version.version_number}"))
    parts.append(_meta_pair("Generado", html.escape(generated_at)))
    parts.append(_meta_pair("Bloques", str(len(blocks))))
    parts.append("</dl>")
    parts.append("</header>")

    if not blocks:
        parts.append("<section class=\"empty\">")
        parts.append(
            "<p>Este reporte no tiene bloques aún. Genera contenido desde el "
            "editor antes de volver a exportar.</p>"
        )
        parts.append("</section>")
    else:
        for index, block in enumerate(blocks, start=1):
            parts.append(_render_block_html(block, index=index))

    parts.append("<footer class=\"footer\">")
    parts.append(
        "<p>Generado por CheckWise · "
        f"{html.escape(generated_at)} · "
        "Powered by Legal Shelf</p>"
    )
    parts.append("</footer>")

    parts.append("</main>")
    parts.append("</body>")
    parts.append("</html>")

    return "\n".join(parts).encode("utf-8")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _audience_label(audience: str | None) -> str:
    if audience == ReportAudience.CLIENT_FACING.value:
        return "Cliente"
    if audience == ReportAudience.VENDOR_FACING.value:
        return "Proveedor"
    if audience == ReportAudience.INTERNAL_ONLY.value:
        return "Interno"
    return audience or "—"


def _meta_pair(label: str, value: str) -> str:
    return (
        f"<div class=\"meta-pair\"><dt>{html.escape(label)}</dt>"
        f"<dd>{value}</dd></div>"
    )


def _render_block_html(block: dict, *, index: int) -> str:
    """Render one block dict into an HTML section.

    Two block types get specialised paths because their natural
    representation is freeform text (``text``) or pure structural
    (``divider``). Everything else uses the generic data-dump path so
    the renderer stays one switch wide — no per-block layouts to
    maintain in 10A.
    """
    block_type = str(block.get("type") or "unknown")
    block_id = str(block.get("id") or f"block-{index}")
    config = block.get("config") or {}
    data = block.get("data") or {}
    ai_summary = block.get("ai_summary") or {}

    if block_type == "divider":
        return "<hr class=\"block-divider\">"

    title = str(config.get("title") or _humanise_type(block_type))
    parts: list[str] = []
    parts.append(f"<section class=\"block\" id=\"{html.escape(block_id)}\">")
    parts.append("<header class=\"block-header\">")
    parts.append(
        f"<span class=\"block-index\">{index:02d}</span>"
        f"<h2>{html.escape(title)}</h2>"
        f"<span class=\"block-type\">{html.escape(block_type)}</span>"
    )
    parts.append("</header>")

    if block_type == "text":
        body = str(config.get("body") or data.get("body") or "")
        if body:
            parts.append("<div class=\"block-text\">")
            for paragraph in body.split("\n\n"):
                paragraph = paragraph.strip()
                if paragraph:
                    parts.append(f"<p>{html.escape(paragraph)}</p>")
            parts.append("</div>")
    else:
        if data:
            parts.append("<div class=\"block-data\">")
            parts.append(_render_data_html(data))
            parts.append("</div>")
        else:
            parts.append(
                "<p class=\"block-empty\">Sin datos para este bloque.</p>"
            )

    ai_text = ai_summary.get("text")
    if ai_text:
        parts.append("<aside class=\"block-ai\">")
        parts.append("<p class=\"ai-label\">Resumen AI</p>")
        for paragraph in str(ai_text).split("\n\n"):
            paragraph = paragraph.strip()
            if paragraph:
                parts.append(f"<p>{html.escape(paragraph)}</p>")
        parts.append("</aside>")

    parts.append("</section>")
    return "".join(parts)


def _humanise_type(block_type: str) -> str:
    return block_type.replace("_", " ").title()


# Plumbing keys that are useful in the live app but pure noise in a printed
# deliverable — dropped from the generic data dump to keep the PDF compact.
_EXPORT_NOISE_KEYS: frozenset[str] = frozenset({
    "id", "vendor_id", "workspace_id", "fetched_at", "as_of", "href",
    "filter_applied", "total_before_filter", "scope_kind", "last_event_at",
    "top", "source_snapshot_id", "data_hash",
})


def _render_data_html(value: Any, *, depth: int = 0) -> str:
    """Recursive data → HTML walker.

    Goal: a structurally faithful dump for arbitrary block payloads.
    Dicts render as definition lists, lists as ordered lists, scalars
    as inline values. Strings are HTML-escaped. Booleans render in
    Spanish ("Sí" / "No"). ``None`` renders as the em-dash so missing
    values are visually distinguishable from empty strings.
    """
    if value is None:
        return "<span class=\"value-null\">—</span>"
    if isinstance(value, bool):
        return (
            "<span class=\"value-bool\">"
            f"{'Sí' if value else 'No'}"
            "</span>"
        )
    if isinstance(value, (int, float)):
        return f"<span class=\"value-num\">{html.escape(str(value))}</span>"
    if isinstance(value, str):
        if not value:
            return "<span class=\"value-empty\">(vacío)</span>"
        return f"<span class=\"value-str\">{html.escape(value)}</span>"
    if isinstance(value, dict):
        # Drop plumbing keys that bloat the export without informing the
        # reader (timestamps, ids, hrefs, scope bookkeeping). They're useful
        # in the live app but noise in a printed deliverable.
        visible = {
            k: v for k, v in value.items() if str(k) not in _EXPORT_NOISE_KEYS
        }
        if not visible:
            return "<span class=\"value-empty\">{}</span>"
        # Leaf dict (all-scalar values, nested) → one compact inline row
        # instead of a stacked dt/dd list. This collapses the bulk of the
        # volume (per-vendor / per-institution entries) from ~7 lines to 1,
        # which is what actually shrinks the printed page count.
        if depth >= 1 and all(
            not isinstance(v, (dict, list, tuple)) for v in visible.values()
        ):
            pairs = " · ".join(
                f"<span class=\"kv\"><b>{html.escape(str(k))}:</b> "
                f"{_render_data_html(v, depth=depth + 1)}</span>"
                for k, v in visible.items()
            )
            return f"<div class=\"data-inline\">{pairs}</div>"
        items = "".join(
            f"<dt>{html.escape(str(k))}</dt>"
            f"<dd>{_render_data_html(v, depth=depth + 1)}</dd>"
            for k, v in visible.items()
        )
        return f"<dl class=\"data-dict depth-{min(depth, 3)}\">{items}</dl>"
    if isinstance(value, (list, tuple)):
        if not value:
            return "<span class=\"value-empty\">(sin entradas)</span>"
        rows = _truncate_iter(value, cap=12)
        items = "".join(
            f"<li>{_render_data_html(item, depth=depth + 1)}</li>" for item in rows
        )
        more = len(value) - len(rows)
        suffix = (
            f"<li class=\"value-empty\">+{more} más…</li>" if more > 0 else ""
        )
        return f"<ol class=\"data-list depth-{min(depth, 3)}\">{items}{suffix}</ol>"
    # Defensive fallback: unknown type → str() it.
    return f"<span class=\"value-other\">{html.escape(str(value))}</span>"


def _truncate_iter(items: Iterable, *, cap: int = 200) -> list:
    """Cap recursive list rendering to keep export HTML reasonable.

    A vendor_risk_matrix block can carry hundreds of vendor entries;
    dumping them all inflates the HTML by 10x with little marginal
    value over a representative sample. 200 keeps small reports
    fully faithful and clips only the genuinely large ones; the
    caller's UI can offer "see all" via the JSON download once
    that lands.
    """
    out = list(items)
    if len(out) > cap:
        out = out[:cap]
        out.append({"_truncated": f"… (limit {cap})"})
    return out


# ---------------------------------------------------------------------------
# Stylesheet (inlined into every export)
# ---------------------------------------------------------------------------


_STYLE_BLOCK = """<style>
:root {
  --color-brand: #013557;
  --color-brand-2: #02558a;
  --color-accent: #09c1b0;
  --color-text: #0f172a;
  --color-text-muted: #475569;
  --color-border: #e2e8f0;
  --color-surface: #ffffff;
  --color-surface-muted: #f8fafc;
}
* { box-sizing: border-box; }
html, body {
  margin: 0;
  padding: 0;
  background: var(--color-surface-muted);
  color: var(--color-text);
  font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", "Open Sans",
    sans-serif;
  font-size: 14px;
  line-height: 1.5;
}
.page {
  max-width: 880px;
  margin: 0 auto;
  padding: 48px 32px;
}
/* Print pagination: drop the screen chrome, tighten spacing, and only keep
   SMALL units (a header with its first rows, a single list item) from being
   split — never whole blocks, which would push tall data blocks onto fresh
   pages and balloon the page count. */
@page { size: Letter; margin: 12mm; }
.data-inline { margin: 2px 0; }
.data-inline .kv { margin-right: 10px; white-space: nowrap; }
.data-inline .kv b { color: var(--color-text-muted); font-weight: 600; }
.block-header { break-after: avoid; page-break-after: avoid; }
.data-list > li, .data-dict > dt, .data-dict > dd { break-inside: avoid; }
@media print {
  html, body { background: #ffffff; font-size: 11px; line-height: 1.35; }
  .page { max-width: none; margin: 0; padding: 0; }
  .cover { margin-bottom: 14px; }
  .cover-meta { margin-bottom: 14px; padding: 8px 0 12px; }
  .block { margin-bottom: 10px; }
  .block-data dl, .block-data ol { margin: 2px 0; }
  .block-data dd { margin: 0 0 2px; }
}
.eyebrow {
  text-transform: uppercase;
  font-size: 11px;
  letter-spacing: 0.12em;
  color: var(--color-text-muted);
  margin: 0 0 8px;
}
.cover h1 {
  font-size: 28px;
  margin: 0 0 12px;
  color: var(--color-brand);
}
.description {
  color: var(--color-text-muted);
  margin: 0 0 24px;
}
.cover-meta {
  display: grid;
  grid-template-columns: repeat(auto-fit, minmax(160px, 1fr));
  gap: 12px;
  padding: 16px 0 24px;
  border-bottom: 1px solid var(--color-border);
  margin-bottom: 32px;
}
.meta-pair {
  display: flex;
  flex-direction: column;
  gap: 2px;
}
.meta-pair dt {
  text-transform: uppercase;
  font-size: 10px;
  letter-spacing: 0.1em;
  color: var(--color-text-muted);
  margin: 0;
}
.meta-pair dd {
  font-size: 13px;
  font-weight: 600;
  margin: 0;
}
.block {
  background: var(--color-surface);
  border: 1px solid var(--color-border);
  border-radius: 8px;
  padding: 20px 24px;
  margin: 0 0 16px;
}
.block-header {
  display: flex;
  align-items: baseline;
  gap: 12px;
  margin-bottom: 16px;
}
.block-index {
  font-family: ui-monospace, SFMono-Regular, "SF Mono", Consolas, monospace;
  font-size: 11px;
  color: var(--color-text-muted);
  border: 1px solid var(--color-border);
  border-radius: 4px;
  padding: 1px 6px;
}
.block-header h2 {
  margin: 0;
  font-size: 16px;
  font-weight: 600;
  color: var(--color-brand);
}
.block-type {
  margin-left: auto;
  font-family: ui-monospace, SFMono-Regular, "SF Mono", Consolas, monospace;
  font-size: 10px;
  color: var(--color-text-muted);
  text-transform: uppercase;
  letter-spacing: 0.06em;
}
.block-text p {
  margin: 0 0 12px;
}
.block-text p:last-child {
  margin-bottom: 0;
}
.block-empty {
  color: var(--color-text-muted);
  font-style: italic;
  margin: 0;
}
.block-divider {
  border: 0;
  border-top: 1px solid var(--color-border);
  margin: 24px 0;
}
.block-data dl {
  display: grid;
  grid-template-columns: minmax(160px, 1fr) 2fr;
  gap: 6px 16px;
  margin: 0;
}
.block-data dl.depth-1, .block-data dl.depth-2, .block-data dl.depth-3 {
  grid-template-columns: minmax(120px, 1fr) 2fr;
  font-size: 13px;
  margin: 4px 0;
}
.block-data dt {
  color: var(--color-text-muted);
  font-weight: 500;
  text-transform: lowercase;
  font-size: 12px;
  letter-spacing: 0.01em;
}
.block-data dd {
  margin: 0;
  word-break: break-word;
}
.block-data ol {
  margin: 4px 0 4px 16px;
  padding: 0;
}
.block-data li {
  margin: 2px 0;
}
.value-num { font-family: ui-monospace, monospace; color: var(--color-brand-2); }
.value-null { color: var(--color-text-muted); }
.value-empty { color: var(--color-text-muted); font-style: italic; }
.value-bool { color: var(--color-accent); font-weight: 600; }
.block-ai {
  margin-top: 16px;
  padding: 12px 16px;
  background: var(--color-surface-muted);
  border-left: 3px solid var(--color-accent);
  border-radius: 4px;
}
.ai-label {
  text-transform: uppercase;
  font-size: 10px;
  letter-spacing: 0.12em;
  color: var(--color-text-muted);
  margin: 0 0 4px;
}
.empty {
  background: var(--color-surface);
  border: 1px dashed var(--color-border);
  border-radius: 8px;
  padding: 32px;
  text-align: center;
  color: var(--color-text-muted);
}
.footer {
  margin-top: 32px;
  padding-top: 16px;
  border-top: 1px solid var(--color-border);
  text-align: center;
  color: var(--color-text-muted);
  font-size: 11px;
}
@media print {
  body { background: var(--color-surface); }
  .page { padding: 16px; }
  .block { page-break-inside: avoid; }
}
</style>"""


__all__ = [
    "SUPPORTED_FORMATS",
    "ReportExportError",
    "ReportExportFormat",
    "render_report_html",
    "render_report_pdf",
    "render_report_xlsx",
    "run_report_export",
    "start_report_export",
]
