"""Helpers for the client-facing metadata workbook."""

from __future__ import annotations

import logging
import os
import re
import tempfile
import unicodedata
import zipfile
from collections import OrderedDict
from dataclasses import dataclass
from pathlib import Path
from xml.etree import ElementTree as ET

from fastapi import HTTPException, status

from app.core.config import settings
from app.models import Client

logger = logging.getLogger(__name__)


@dataclass(frozen=True)
class MetadataSheetPreview:
    name: str
    rows: list[list[str]]


# perf audit P0-1 — the master workbook was unzipped + XML-parsed synchronously
# on every /client/metadata (and /admin) load, re-parsing identical bytes each
# time. Cache the parsed preview keyed by the file's identity (path + mtime +
# size): a re-upload changes mtime/size and busts the entry automatically. The
# cache is intentionally tiny and process-local — a bounded LRU, no TTL needed.
_PREVIEW_CACHE_MAX_ENTRIES = 32
_preview_cache: OrderedDict[
    tuple[str, int, int, int, int], list[MetadataSheetPreview]
] = OrderedDict()


def client_master_file_path(client: Client) -> Path:
    return (
        Path(settings.METADATA_EXPORT_PATH).expanduser().resolve()
        / _export_slug(client.name)
        / "client_master_metadata.xlsx"
    )


def display_export_path(value: object) -> str | None:
    if not isinstance(value, str) or not value.strip():
        return None
    path = Path(value).expanduser().resolve()
    export_root = Path(settings.METADATA_EXPORT_PATH).expanduser().resolve()
    try:
        return str(path.relative_to(export_root))
    except ValueError:
        return path.name


def read_xlsx_preview(
    path: Path, *, max_rows_per_sheet: int = 40, max_columns: int = 12
) -> list[MetadataSheetPreview]:
    try:
        stat = path.stat()
        cache_key = (
            str(path),
            stat.st_mtime_ns,
            stat.st_size,
            max_rows_per_sheet,
            max_columns,
        )
        cached = _preview_cache.get(cache_key)
        if cached is not None:
            _preview_cache.move_to_end(cache_key)  # LRU bump
            return cached
        previews = _parse_xlsx_preview(
            path, max_rows_per_sheet=max_rows_per_sheet, max_columns=max_columns
        )
        _preview_cache[cache_key] = previews
        _preview_cache.move_to_end(cache_key)
        while len(_preview_cache) > _PREVIEW_CACHE_MAX_ENTRIES:
            _preview_cache.popitem(last=False)
        return previews
    except (KeyError, zipfile.BadZipFile, ET.ParseError, OSError) as exc:
        # P3-6 — never echo parser internals (zip/XML/path fragments) across the
        # trust boundary; the workbook is server-generated, so the client only
        # needs a static message. The detail is logged server-side instead.
        logger.warning("metadata xlsx preview failed for %s: %s", path, exc)
        raise HTTPException(
            status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="No se pudo leer el archivo de metadata.",
        ) from exc


def _parse_xlsx_preview(
    path: Path, *, max_rows_per_sheet: int, max_columns: int
) -> list[MetadataSheetPreview]:
    with zipfile.ZipFile(path, "r") as archive:
        workbook_xml = archive.read("xl/workbook.xml")
        sheet_names = _xlsx_sheet_names(workbook_xml)
        previews = []
        for index, name in enumerate(sheet_names, start=1):
            worksheet_path = f"xl/worksheets/sheet{index}.xml"
            if worksheet_path not in archive.namelist():
                continue
            previews.append(
                MetadataSheetPreview(
                    name=name,
                    rows=_xlsx_sheet_rows(
                        archive.read(worksheet_path),
                        max_rows=max_rows_per_sheet,
                        max_columns=max_columns,
                    ),
                )
            )
        return previews


def _export_slug(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value or "")
    ascii_text = normalized.encode("ascii", "ignore").decode("ascii")
    clean = re.sub(
        r"\s+",
        " ",
        re.sub(r"[^a-zA-Z0-9]+", " ", ascii_text).lower(),
    ).strip()
    return re.sub(r"[^a-z0-9-]+", "-", clean.replace(" ", "-")).strip("-") or "unknown"


def _xlsx_sheet_names(workbook_xml: bytes) -> list[str]:
    namespace = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    root = ET.fromstring(workbook_xml)
    return [
        sheet.attrib.get("name", f"Sheet {index}")
        for index, sheet in enumerate(root.findall(".//x:sheet", namespace), start=1)
    ]


def _xlsx_sheet_rows(
    worksheet_xml: bytes, *, max_rows: int, max_columns: int
) -> list[list[str]]:
    namespace = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    root = ET.fromstring(worksheet_xml)
    parsed_rows: list[list[str]] = []
    for row in root.findall(".//x:sheetData/x:row", namespace)[:max_rows]:
        values = [""] * max_columns
        for cell in row.findall("x:c", namespace):
            column_index = _xlsx_column_index(cell.attrib.get("r", "A1")) - 1
            if not 0 <= column_index < max_columns:
                continue
            values[column_index] = _xlsx_cell_text(cell, namespace)
        while values and values[-1] == "":
            values.pop()
        parsed_rows.append(values)
    return parsed_rows


def _xlsx_column_index(cell_ref: str) -> int:
    index = 0
    for char in cell_ref:
        if not char.isalpha():
            break
        index = index * 26 + (ord(char.upper()) - 64)
    return max(index, 1)


def _xlsx_cell_text(cell: ET.Element, namespace: dict[str, str]) -> str:
    inline = cell.find("x:is", namespace)
    if inline is not None:
        return "".join(node.text or "" for node in inline.findall(".//x:t", namespace))
    value = cell.find("x:v", namespace)
    return value.text if value is not None and value.text is not None else ""


def filter_master_by_vendor(
    master_path: Path, *, vendor_name: str, period_key: str | None = None
) -> Path | None:
    """Write a temp copy of the client master keeping only one provider's rows.

    CW-15 — the all-providers master already exists; this is the per-provider
    (optionally per-period) view. Reuses the single master workbook (one R2
    fetch) and drops every "01 Metadata" row whose ``Proveedor`` (and, when
    given, ``Periodo``) doesn't match, preserving the workbook structure and
    the "00 Guia" sheet. Returns the temp path, or ``None`` when the provider
    has no rows in the master.
    """
    from openpyxl import load_workbook

    workbook = load_workbook(master_path)
    if "01 Metadata" not in workbook.sheetnames:
        return None
    sheet = workbook["01 Metadata"]
    header = [str(cell.value or "").strip() for cell in sheet[1]]
    if "Proveedor" not in header:
        return None
    vendor_col = header.index("Proveedor") + 1
    period_col = (
        header.index("Periodo") + 1 if (period_key and "Periodo" in header) else None
    )
    wanted_vendor = (vendor_name or "").strip()
    wanted_period = (period_key or "").strip()

    kept = 0
    # Bottom-up so deletions don't reindex rows we haven't checked yet.
    for row in range(sheet.max_row, 1, -1):
        vendor_cell = str(sheet.cell(row=row, column=vendor_col).value or "").strip()
        keep = vendor_cell == wanted_vendor
        if keep and period_col is not None:
            period_cell = str(
                sheet.cell(row=row, column=period_col).value or ""
            ).strip()
            keep = period_cell == wanted_period
        if keep:
            kept += 1
        else:
            sheet.delete_rows(row, 1)
    if kept == 0:
        return None

    handle, temp_name = tempfile.mkstemp(prefix="cw-vendor-metadata-", suffix=".xlsx")
    os.close(handle)
    output = Path(temp_name)
    workbook.save(output)
    return output
