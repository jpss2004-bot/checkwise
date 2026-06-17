"""CW-15 — per-provider metadata export (filter the client master by vendor)."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.services.client_metadata import filter_master_by_vendor

_HEADER = ["Cliente", "Proveedor", "Periodo", "Nombre del documento"]


def _make_master(tmp_path: Path) -> Path:
    """A minimal client master with the real sheet names + two providers."""
    wb = Workbook()
    guia = wb.active
    guia.title = "00 Guia"
    guia["A1"] = "Guía"
    meta = wb.create_sheet("01 Metadata")
    meta.append(_HEADER)
    meta.append(["Cliente X", "Proveedor A", "2026-M01", "doc a1"])
    meta.append(["Cliente X", "Proveedor B", "2026-M01", "doc b1"])
    meta.append(["Cliente X", "Proveedor A", "2026-M02", "doc a2"])
    path = tmp_path / "client_master_metadata.xlsx"
    wb.save(path)
    return path


def test_filter_keeps_only_the_target_provider(tmp_path: Path) -> None:
    out = filter_master_by_vendor(_make_master(tmp_path), vendor_name="Proveedor A")
    assert out is not None
    try:
        wb = load_workbook(out)
        # Structure preserved (guía sheet survives).
        assert "00 Guia" in wb.sheetnames
        rows = list(wb["01 Metadata"].iter_rows(values_only=True))
        assert list(rows[0]) == _HEADER  # header kept
        assert {r[1] for r in rows[1:]} == {"Proveedor A"}  # only A's rows
        assert len(rows) == 3  # header + 2 Proveedor A rows
    finally:
        out.unlink(missing_ok=True)


def test_filter_by_vendor_and_period(tmp_path: Path) -> None:
    out = filter_master_by_vendor(
        _make_master(tmp_path), vendor_name="Proveedor A", period_key="2026-M02"
    )
    assert out is not None
    try:
        rows = list(load_workbook(out)["01 Metadata"].iter_rows(values_only=True))
        assert len(rows) == 2  # header + the single A / 2026-M02 row
        assert rows[1][3] == "doc a2"
    finally:
        out.unlink(missing_ok=True)


def test_filter_unknown_provider_returns_none(tmp_path: Path) -> None:
    assert (
        filter_master_by_vendor(_make_master(tmp_path), vendor_name="Nadie") is None
    )
